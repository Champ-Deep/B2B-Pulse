import csv
import io
import logging
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_user
from app.database import get_db
from app.models.post import Post
from app.models.tracked_page import PollingMode, TrackedPage, TrackedPageSubscription
from app.models.user import User
from app.schemas.tracked_page import (
    EngagementBrief,
    ImportResult,
    PostSubmitRequest,
    PostWithEngagements,
    SubscriptionCreate,
    SubscriptionResponse,
    TrackedPageCreate,
    TrackedPageResponse,
    TrackedPageUpdate,
)
from app.services.url_utils import (
    detect_page_type,
    detect_platform,
    extract_external_id,
    extract_post_id,
    normalize_url,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/tracked-pages", tags=["tracked-pages"])


@router.post("", response_model=TrackedPageResponse, status_code=status.HTTP_201_CREATED, summary="Create Tracked Page")
async def create_tracked_page(
    request: TrackedPageCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add a new social media page to track for auto-engagement."""
    try:
        platform = detect_platform(request.url)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported platform for URL: {request.url}",
        ) from exc
    normalized = normalize_url(request.url)

    # Check for duplicate URL in org
    existing = await db.execute(
        select(TrackedPage).where(
            TrackedPage.org_id == current_user.org_id,
            TrackedPage.url == normalized,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="This page is already being tracked")

    external_id = extract_external_id(request.url, platform)
    page_type = detect_page_type(request.url, platform)

    page = TrackedPage(
        org_id=current_user.org_id,
        platform=platform,
        external_id=external_id,
        url=normalized,
        name=request.name or external_id or request.url,
        page_type=page_type,
    )
    db.add(page)
    await db.flush()

    # Auto-subscribe ALL active org members
    all_users_result = await db.execute(
        select(User).where(
            User.org_id == current_user.org_id,
            User.is_active.is_(True),
        )
    )
    for member in all_users_result.scalars().all():
        db.add(
            TrackedPageSubscription(
                tracked_page_id=page.id,
                user_id=member.id,
                auto_like=True,
                auto_comment=True,
                polling_mode=PollingMode.NORMAL,
            )
        )
    await db.flush()

    return page


@router.get("", response_model=list[TrackedPageResponse], summary="List Tracked Pages")
async def list_tracked_pages(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all tracked pages for the current user's organization."""
    result = await db.execute(
        select(TrackedPage)
        .where(TrackedPage.org_id == current_user.org_id)
        .order_by(TrackedPage.created_at.desc())
    )
    return result.scalars().all()


@router.put("/{page_id}", response_model=TrackedPageResponse, summary="Update Tracked Page")
async def update_tracked_page(
    page_id: uuid.UUID,
    request: TrackedPageUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update a tracked page's name or active status."""
    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    page = result.scalar_one_or_none()
    if not page:
        raise HTTPException(status_code=404, detail="Tracked page not found")

    if request.name is not None:
        page.name = request.name
    if request.active is not None:
        page.active = request.active
    return page


@router.delete("/{page_id}", status_code=status.HTTP_204_NO_CONTENT, summary="Delete Tracked Page")
async def delete_tracked_page(
    page_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a tracked page and its subscriptions."""
    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    page = result.scalar_one_or_none()
    if not page:
        raise HTTPException(status_code=404, detail="Tracked page not found")
    await db.delete(page)


@router.post("/{page_id}/subscribe", response_model=SubscriptionResponse, status_code=201, summary="Subscribe to Tracked Page")
async def subscribe_to_page(
    page_id: uuid.UUID,
    request: SubscriptionCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Subscribe the current user to a tracked page for auto-engagement."""
    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Tracked page not found")

    result = await db.execute(
        select(TrackedPageSubscription).where(
            TrackedPageSubscription.tracked_page_id == page_id,
            TrackedPageSubscription.user_id == current_user.id,
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Already subscribed")

    sub = TrackedPageSubscription(
        tracked_page_id=page_id,
        user_id=current_user.id,
        auto_like=request.auto_like,
        auto_comment=request.auto_comment,
        polling_mode=PollingMode(request.polling_mode),
        tags=request.tags,
    )
    db.add(sub)
    await db.flush()
    return sub


@router.put("/{page_id}/subscribe", response_model=SubscriptionResponse, summary="Update Page Subscription")
async def update_subscription(
    page_id: uuid.UUID,
    request: SubscriptionCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update subscription settings for a tracked page."""
    result = await db.execute(
        select(TrackedPageSubscription).where(
            TrackedPageSubscription.tracked_page_id == page_id,
            TrackedPageSubscription.user_id == current_user.id,
        )
    )
    sub = result.scalar_one_or_none()
    if not sub:
        raise HTTPException(status_code=404, detail="Subscription not found")

    sub.auto_like = request.auto_like
    sub.auto_comment = request.auto_comment
    sub.polling_mode = PollingMode(request.polling_mode)
    sub.tags = request.tags
    return sub


# --- CSV/Excel Bulk Import ---


@router.post("/import", response_model=ImportResult, summary="Bulk Import Tracked Pages")
async def import_tracked_pages(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Import tracked pages from a CSV or Excel file.

    CSV format: url,name (header row required)
    Excel format: .xlsx with 'url' and optional 'name' columns
    """
    content = await file.read()
    rows: list[dict[str, str]] = []

    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            if not ws:
                raise HTTPException(status_code=400, detail="Empty workbook")
            headers = [
                str(cell.value or "").strip().lower()
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            if "url" not in headers:
                raise HTTPException(status_code=400, detail="Missing 'url' column in Excel file")
            url_idx = headers.index("url")
            name_idx = headers.index("name") if "name" in headers else None
            for row in ws.iter_rows(min_row=2, values_only=True):
                url_val = str(row[url_idx] or "").strip()
                name_val = (
                    str(row[name_idx] or "").strip()
                    if name_idx is not None and row[name_idx]
                    else ""
                )
                if url_val:
                    rows.append({"url": url_val, "name": name_val})
        except ImportError as exc:
            raise HTTPException(
                status_code=400, detail="Excel support requires openpyxl. Please use CSV format."
            ) from exc
    else:
        # Assume CSV
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames or "url" not in [f.lower().strip() for f in reader.fieldnames]:
            raise HTTPException(status_code=400, detail="CSV must have a 'url' column header")
        for row in reader:
            # Handle case-insensitive column names
            url_val = (row.get("url") or row.get("URL") or row.get("Url") or "").strip()
            name_val = (row.get("name") or row.get("Name") or row.get("NAME") or "").strip()
            if url_val:
                rows.append({"url": url_val, "name": name_val})

    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows):
        url = normalize_url(row["url"])
        name = row.get("name", "")

        try:
            platform = detect_platform(url)
        except ValueError:
            platform = None
        if not platform:
            errors.append(f"Row {i + 2}: Unsupported platform for URL: {url}")
            continue

        # Check for duplicate URL in org
        existing = await db.execute(
            select(TrackedPage).where(
                TrackedPage.org_id == current_user.org_id,
                TrackedPage.url == url,
            )
        )
        if existing.scalar_one_or_none():
            skipped += 1
            continue

        external_id = extract_external_id(url, platform)
        page_type = detect_page_type(url, platform)

        page = TrackedPage(
            org_id=current_user.org_id,
            platform=platform,
            external_id=external_id,
            url=url,
            name=name or external_id or url,
            page_type=page_type,
        )
        db.add(page)
        await db.flush()

        # Auto-subscribe ALL active org members
        members_result = await db.execute(
            select(User).where(
                User.org_id == current_user.org_id,
                User.is_active.is_(True),
            )
        )
        for member in members_result.scalars().all():
            db.add(
                TrackedPageSubscription(
                    tracked_page_id=page.id,
                    user_id=member.id,
                    auto_like=True,
                    auto_comment=True,
                    polling_mode=PollingMode.NORMAL,
                )
            )
        imported += 1

    await db.flush()
    return ImportResult(imported=imported, skipped=skipped, errors=errors)


# --- Manual Post URL Submission ---


@router.post("/{page_id}/submit-post", status_code=201, summary="Submit Post URL")
async def submit_post_url(
    page_id: uuid.UUID,
    request: PostSubmitRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Manually submit a post URL to trigger auto-engagement."""
    # Verify page belongs to user's org
    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    page = result.scalar_one_or_none()
    if not page:
        raise HTTPException(status_code=404, detail="Tracked page not found")

    # Extract external post ID based on platform
    external_post_id = extract_post_id(request.url, page.platform)
    if not external_post_id:
        raise HTTPException(
            status_code=400,
            detail="Could not extract post ID from URL. Please check the URL format.",
        )

    # Deduplicate
    existing = await db.execute(select(Post).where(Post.external_post_id == external_post_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Post already submitted")

    # Create post
    post = Post(
        tracked_page_id=page.id,
        platform=page.platform,
        external_post_id=external_post_id,
        url=request.url,
    )
    db.add(post)
    await db.flush()

    # Enqueue engagement
    from app.workers.engagement_tasks import schedule_staggered_engagements

    schedule_staggered_engagements.delay(str(post.id), str(page.id))

    logger.info(f"Manual post submitted: {request.url} for page {page.name}")
    return {"message": "Post submitted for engagement", "post_id": str(post.id)}


# --- Poll Now ---


@router.post("/{page_id}/poll-now", summary="Poll Page Now")
async def poll_page_now(
    page_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Trigger an immediate poll for a tracked page."""
    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    page = result.scalar_one_or_none()
    if not page:
        raise HTTPException(status_code=404, detail="Tracked page not found")
    if not page.active:
        raise HTTPException(status_code=400, detail="Page is inactive — activate it first")

    from app.workers.polling_tasks import poll_single_page_task

    poll_single_page_task.delay(str(page.id))
    return {"message": "Poll triggered", "page_id": str(page.id)}


# --- Page Posts with Engagement Status ---


@router.get("/{page_id}/posts", response_model=list[PostWithEngagements], summary="Get Page Posts")
async def get_page_posts(
    page_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    limit: int = 20,
):
    """Get recent posts for a tracked page with their engagement status."""
    from app.models.engagement import EngagementAction

    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Tracked page not found")

    posts_result = await db.execute(
        select(Post)
        .where(Post.tracked_page_id == page_id)
        .order_by(Post.first_seen_at.desc())
        .limit(limit)
    )
    posts = posts_result.scalars().all()

    response = []
    for post in posts:
        eng_result = await db.execute(
            select(EngagementAction).where(EngagementAction.post_id == post.id)
        )
        engagements = eng_result.scalars().all()

        response.append(
            PostWithEngagements(
                id=post.id,
                url=post.url,
                content_text=post.content_text,
                external_post_id=post.external_post_id,
                first_seen_at=post.first_seen_at,
                engagements=[
                    EngagementBrief(
                        id=e.id,
                        action_type=e.action_type.value,
                        status=e.status.value,
                        completed_at=e.completed_at,
                        error_message=e.error_message,
                    )
                    for e in engagements
                ],
            )
        )

    return response


# --- Poll Status ---


@router.get("/{page_id}/poll-status", summary="Get Poll Status")
async def get_poll_status(
    page_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get the last poll status for a tracked page.

    Reads from Redis first (fast path). Falls back to the DB columns
    `last_polled_at` / `last_poll_status` when the Redis key has expired.
    """
    import json

    import redis as sync_redis

    from app.config import settings

    result = await db.execute(
        select(TrackedPage).where(
            TrackedPage.id == page_id, TrackedPage.org_id == current_user.org_id
        )
    )
    page = result.scalar_one_or_none()
    if not page:
        raise HTTPException(status_code=404, detail="Tracked page not found")

    # Fast path: Redis
    r = sync_redis.from_url(settings.redis_url)
    raw = r.get(f"autoengage:poll_status:{page_id}")
    if raw:
        return json.loads(raw)

    # Fallback: DB persistent columns
    if page.last_polled_at:
        return {
            "status": page.last_poll_status or "ok",
            "last_polled_at": page.last_polled_at.isoformat(),
            "posts_found": None,
            "new_posts": None,
            "error": None,
        }

    return {"status": "never_polled"}

